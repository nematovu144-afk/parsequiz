"""Unit tests for JSON/XLSX/CSV export serializers."""

import csv
import io
import json

from openpyxl import load_workbook

from app.export.exporters import export_csv, export_json, export_xlsx
from app.schemas.quiz import Question

QUESTIONS = [
    Question(
        question="Capital of France?",
        options=["London", "Paris", "Berlin"],
        correct_option_index=1,
        explanation="Paris has been the capital since 987.",
    ),
]


def test_export_json_roundtrip():
    data = json.loads(export_json(QUESTIONS))
    assert len(data) == 1
    assert data[0]["question"] == "Capital of France?"
    assert data[0]["options"] == ["London", "Paris", "Berlin"]
    assert data[0]["correct_option_index"] == 1
    # internal validation flags must not leak into the export
    assert "flags" not in data[0]


def test_export_csv_has_header_and_correct_letter():
    text = export_csv(QUESTIONS).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0][:2] == ["#", "Question"]
    assert rows[1][1] == "Capital of France?"
    assert rows[1][2:5] == ["London", "Paris", "Berlin"]
    assert rows[1][5] == "B"  # correct_option_index=1 -> letter B (3 options -> Correct at col 5)


def test_export_xlsx_has_correct_letter_and_headers():
    wb = load_workbook(io.BytesIO(export_xlsx(QUESTIONS)))
    ws = wb.active
    assert ws.cell(row=1, column=2).value == "Question"
    assert ws.cell(row=2, column=2).value == "Capital of France?"
    assert ws.cell(row=2, column=6).value == "B"  # 3 options -> Correct at col 6


def test_export_empty_list_does_not_crash():
    assert json.loads(export_json([])) == []
    export_csv([])
    export_xlsx([])


# ── Mixed option counts (some questions 3 options, some 6) ──────


MIXED_QUESTIONS = [
    Question(question="2+2?", options=["3", "4"], correct_option_index=1),
    Question(
        question="Pick a color",
        options=["Red", "Green", "Blue", "Yellow", "Purple", "Orange"],
        correct_option_index=5,
    ),
]


def test_export_csv_sizes_columns_to_widest_question():
    text = export_csv(MIXED_QUESTIONS).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0] == ["#", "Question", "Option A", "Option B", "Option C",
                        "Option D", "Option E", "Option F", "Correct", "Explanation"]
    # 6-option question keeps all six options, nothing truncated
    assert rows[2][2:8] == ["Red", "Green", "Blue", "Yellow", "Purple", "Orange"]
    assert rows[2][8] == "F"  # correct_option_index=5 -> letter F


def test_export_xlsx_sizes_columns_to_widest_question():
    wb = load_workbook(io.BytesIO(export_xlsx(MIXED_QUESTIONS)))
    ws = wb.active
    assert ws.cell(row=1, column=8).value == "Option F"
    assert ws.cell(row=3, column=8).value == "Orange"  # 6th option not dropped
    assert ws.cell(row=3, column=9).value == "F"  # Correct column shifted past 6 options


# ── Formula-injection hardening (CWE-1236) ──────────────────────

INJECTION_QUESTIONS = [
    Question(
        question="=cmd|'/c calc'!A1",
        options=["+SUM(A1:A9)", "-1+1", "@SUM(1,1)", "safe option"],
        correct_option_index=3,
        explanation="=HYPERLINK(\"http://evil.example\")",
    ),
]


def test_export_csv_neutralises_formula_prefixes():
    text = export_csv(INJECTION_QUESTIONS).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    question, opt_a, opt_b, opt_c, opt_d, _correct, explanation = rows[1][1:8]
    assert question.startswith("'=")
    assert opt_a.startswith("'+")
    assert opt_b.startswith("'-")
    assert opt_c.startswith("'@")
    assert opt_d == "safe option"  # untouched — no leading trigger char
    assert explanation.startswith("'=")


def test_export_xlsx_neutralises_formula_prefixes():
    wb = load_workbook(io.BytesIO(export_xlsx(INJECTION_QUESTIONS)))
    ws = wb.active
    assert ws.cell(row=2, column=2).value.startswith("'=")
    assert ws.cell(row=2, column=3).value.startswith("'+")
    assert ws.cell(row=2, column=6).value == "safe option"
