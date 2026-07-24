"""Serialise Question lists into downloadable formats."""

from __future__ import annotations

import csv
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.quiz import Question

# Leading characters that Excel/Google Sheets/LibreOffice interpret as the
# start of a formula. A cell value beginning with one of these, if opened by
# a spreadsheet app, can execute arbitrary formulas (CWE-1236) — since these
# strings originate from uploaded documents, they're untrusted.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value: str) -> str:
    """Neutralise formula-injection payloads for spreadsheet exports."""
    if value and value[0] in _FORMULA_TRIGGER_CHARS:
        return "'" + value
    return value


def export_json(questions: list[Question]) -> bytes:
    """Clean JSON array (no internal flags)."""
    payload = [
        {
            "id": str(q.id),
            "question": q.question,
            "options": q.options,
            "correct_option_index": q.correct_option_index,
            "explanation": q.explanation,
        }
        for q in questions
    ]
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_xlsx(questions: list[Question]) -> bytes:
    """Styled Excel workbook with one row per question.

    Option columns are sized to the widest question — a 3-option and a
    6-option question in the same set both export in full, not just A-D.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz"

    max_opts = max((len(q.options) for q in questions), default=2)
    option_headers = [f"Option {chr(65 + i)}" for i in range(max_opts)]
    headers = ["#", "Question", *option_headers, "Correct", "Explanation"]
    correct_col = 3 + max_opts
    explanation_col = correct_col + 1

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, q in enumerate(questions, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=_sanitize_cell(q.question))
        for j, opt in enumerate(q.options):
            ws.cell(row=i, column=3 + j, value=_sanitize_cell(opt))
        if q.correct_option_index is not None and q.correct_option_index < len(q.options):
            letter = chr(65 + q.correct_option_index)  # A, B, C, ...
            ws.cell(row=i, column=correct_col, value=letter)
        ws.cell(row=i, column=explanation_col, value=_sanitize_cell(q.explanation or ""))

    # Column widths
    ws.column_dimensions["B"].width = 50
    for col in range(3, 3 + max_opts):
        ws.column_dimensions[get_column_letter(col)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(questions: list[Question]) -> bytes:
    """UTF-8 CSV with BOM for Excel compatibility.

    Option columns are sized to the widest question, same as export_xlsx.
    """
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM
    writer = csv.writer(buf)

    max_opts = max((len(q.options) for q in questions), default=2)
    option_headers = [f"Option {chr(65 + i)}" for i in range(max_opts)]
    writer.writerow(["#", "Question", *option_headers, "Correct", "Explanation"])

    for i, q in enumerate(questions, 1):
        opts = q.options + [""] * (max_opts - len(q.options))  # pad to max_opts
        correct = ""
        if q.correct_option_index is not None and q.correct_option_index < len(q.options):
            correct = chr(65 + q.correct_option_index)
        writer.writerow([
            i,
            _sanitize_cell(q.question),
            *[_sanitize_cell(o) for o in opts],
            correct,
            _sanitize_cell(q.explanation or ""),
        ])

    return buf.getvalue().encode("utf-8")
